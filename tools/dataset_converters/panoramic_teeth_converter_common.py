from __future__ import annotations

import ast
import random
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from PIL import Image


TARGET_TEETH = [
    11, 12, 13, 14, 15, 16, 17, 18,
    21, 22, 23, 24, 25, 26, 27, 28,
    31, 32, 33, 34, 35, 36, 37, 38,
    41, 42, 43, 44, 45, 46, 47, 48,
]
VALID_TOOTH_SET = set(TARGET_TEETH)
VIS_VISIBLE = 2
VIS_MISSING = 0

MANUAL_LABEL_RE = re.compile(r'^(?P<tooth>\d+)(?P<side>M|D|近中|远中)$')
AUTO_LABEL_RE = re.compile(r'^auto_(?P<tooth>-?\d+)_(?P<side>left|right)$')


@dataclass
class RawEntry:
    tooth_id: Optional[int]
    side_token: Optional[str]
    kind: str
    tooth_id_from_label: bool
    side_from_label: bool
    raw_points: List[Tuple[float, float]]
    label: str
    row_index: int


@dataclass
class ImageRecord:
    sample_id: int
    image_path: Path
    traced_image_path: Optional[Path]
    width: int
    height: int
    side_entries: Dict[int, List[RawEntry]]


def detect_subdir(root: Path, keywords: Sequence[str]) -> Path:
    for child in sorted(root.iterdir()):
        if child.is_dir() and any(keyword in child.name for keyword in keywords):
            return child
    raise FileNotFoundError(
        f'Unable to detect subdirectory under {root} with keywords '
        f'{keywords!r}.')


def extract_sample_id(path: Path) -> int:
    match = re.search(r'(\d+)', path.stem)
    if not match:
        raise ValueError(f'Unable to extract sample id from {path.name}')
    return int(match.group(1))


def normalize_tooth_id(value: object) -> Optional[int]:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        tooth_id = int(value)
        return tooth_id if tooth_id in VALID_TOOTH_SET else None

    text = str(value).strip()
    if not text:
        return None

    match = re.match(r'\s*(-?\d+(?:\.\d+)?)', text)
    if not match:
        return None

    tooth_id = int(float(match.group(1)))
    return tooth_id if tooth_id in VALID_TOOTH_SET else None


def is_same_point(point_a: Tuple[float, float],
                  point_b: Tuple[float, float],
                  eps: float = 1e-6) -> bool:
    return (abs(point_a[0] - point_b[0]) <= eps and
            abs(point_a[1] - point_b[1]) <= eps)


def point_or_missing(point: Optional[Tuple[float, float]],
                     visible: bool) -> Tuple[List[float], int]:
    if point is None or not visible:
        return [0.0, 0.0], VIS_MISSING
    return [float(point[0]), float(point[1])], VIS_VISIBLE


def resolve_source_canvas_width(image_width: int,
                                image_height: int,
                                source_canvas_width: Optional[int],
                                source_canvas_height: int) -> int:
    if source_canvas_width is not None:
        return source_canvas_width
    return max(1, round(image_width * source_canvas_height / float(image_height)))


def scale_points_to_image(points: Sequence[Tuple[float, float]],
                          image_width: int,
                          image_height: int,
                          source_canvas_width: Optional[int],
                          source_canvas_height: int,
                          point_offset_x: float = 0.0,
                          point_offset_y: float = 0.0
                          ) -> List[Tuple[float, float]]:
    resolved_canvas_width = resolve_source_canvas_width(
        image_width,
        image_height,
        source_canvas_width,
        source_canvas_height,
    )
    scale_x = image_width / float(resolved_canvas_width)
    scale_y = image_height / float(source_canvas_height)
    return [
        (point[0] * scale_x + point_offset_x, point[1] * scale_y + point_offset_y)
        for point in points
    ]


def find_measurement_text(row: Sequence[object]) -> Optional[str]:
    for value in row:
        if isinstance(value, str) and 'PointList:' in value and 'label_info:' in value:
            return value
    return None


def split_measurement_segments(text: str) -> Iterable[str]:
    start_positions = [match.start() for match in re.finditer(r'label_info:', text)]
    for index, start in enumerate(start_positions):
        end = start_positions[index + 1] if index + 1 < len(start_positions) else len(text)
        segment = text[start:end].strip()
        if segment:
            yield segment


def extract_list_literal(segment: str) -> Optional[str]:
    marker_index = segment.find('PointList:')
    if marker_index < 0:
        return None
    tail = segment[marker_index + len('PointList:'):]
    left_index = tail.find('[')
    if left_index < 0:
        return None

    depth = 0
    start = None
    for index, char in enumerate(tail[left_index:], start=left_index):
        if char == '[':
            if depth == 0:
                start = index
            depth += 1
        elif char == ']':
            depth -= 1
            if depth == 0 and start is not None:
                return tail[start:index + 1]
    return None


def normalize_label(label: str) -> str:
    return ''.join(label.split())


def parse_label(label: str) -> Tuple[Optional[int], Optional[str], str]:
    label = normalize_label(label)
    if not label or label.lower() == 'null':
        return None, None, 'null'

    auto_match = AUTO_LABEL_RE.fullmatch(label)
    if auto_match:
        return int(auto_match.group('tooth')), auto_match.group('side'), 'auto'

    manual_match = MANUAL_LABEL_RE.fullmatch(label)
    if manual_match:
        side = manual_match.group('side')
        if side == '近中':
            side = 'M'
        elif side == '远中':
            side = 'D'
        return int(manual_match.group('tooth')), side, 'manual'

    return None, None, 'unknown'


def parse_points(list_literal: str) -> List[Tuple[float, float]]:
    parsed = ast.literal_eval(list_literal)
    return [(float(point['x']), float(point['y'])) for point in parsed]


def parse_measurement_entries(text: str,
                              row_index: int,
                              fallback_tooth_id: Optional[int],
                              summary: dict) -> List[RawEntry]:
    entries: List[RawEntry] = []
    for segment in split_measurement_segments(text):
        try:
            label_part = segment.split('message:', 1)[0].split(
                'label_info:', 1)[1].strip()
        except IndexError:
            summary['segment_without_label'] += 1
            continue

        list_literal = extract_list_literal(segment)
        if list_literal is None:
            summary['segment_without_pointlist'] += 1
            continue

        try:
            raw_points = parse_points(list_literal)
        except Exception:
            summary['pointlist_parse_error'] += 1
            continue

        tooth_id, side_token, kind = parse_label(label_part)
        tooth_id_from_label = tooth_id is not None
        side_from_label = side_token is not None
        if tooth_id is None:
            tooth_id = fallback_tooth_id

        entries.append(
            RawEntry(
                tooth_id=tooth_id,
                side_token=side_token,
                kind=kind,
                tooth_id_from_label=tooth_id_from_label,
                side_from_label=side_from_label,
                raw_points=raw_points,
                label=normalize_label(label_part),
                row_index=row_index,
            ))
    return entries


def candidate_quality(entry: RawEntry) -> Tuple[int, int, int]:
    dedup_count = 0
    previous: Optional[Tuple[float, float]] = None
    for point in entry.raw_points:
        if previous is None or not is_same_point(previous, point):
            dedup_count += 1
            previous = point
    explicit = 1 if entry.side_token in {'M', 'D'} else 0
    return explicit, dedup_count, len(entry.raw_points)


def mesial_is_image_right(tooth_id: int) -> bool:
    return tooth_id // 10 in {1, 4}


def image_side_to_md(tooth_id: int, image_side: str) -> str:
    if mesial_is_image_right(tooth_id):
        return 'M' if image_side == 'right' else 'D'
    return 'M' if image_side == 'left' else 'D'


def convert_manual_side_points(raw_points: Sequence[Tuple[float, float]]
                               ) -> Tuple[List[List[float]], List[int]]:
    main_points: List[Optional[Tuple[float, float]]] = [None] * 5
    visibilities: List[int] = [VIS_MISSING] * 5
    main_indices = [0, 1, 3, 4, 5] if len(raw_points) >= 6 else [0, 1, 2, 3, 4]

    for target_index, source_index in enumerate(main_indices):
        if source_index < len(raw_points):
            main_points[target_index] = raw_points[source_index]
            visibilities[target_index] = VIS_VISIBLE

    al_point = None
    al_visible = False
    if len(raw_points) >= 6:
        candidate = raw_points[2]
        if not is_same_point(candidate, raw_points[1]):
            al_point = candidate
            al_visible = True

    converted_points: List[List[float]] = []
    normalized_visibilities: List[int] = []
    for point, visibility in zip(main_points, visibilities):
        xy, vis = point_or_missing(point, visibility == VIS_VISIBLE)
        converted_points.append(xy)
        normalized_visibilities.append(vis)

    al_xy, al_vis = point_or_missing(al_point, al_visible)
    converted_points.append(al_xy)
    normalized_visibilities.append(al_vis)
    return converted_points, normalized_visibilities


def convert_auto_side_points(raw_points: Sequence[Tuple[float, float]]
                             ) -> Tuple[List[List[float]], List[int]]:
    converted_points: List[List[float]] = []
    visibilities: List[int] = []
    for index in range(5):
        point = raw_points[index] if index < len(raw_points) else None
        xy, vis = point_or_missing(point, point is not None)
        converted_points.append(xy)
        visibilities.append(vis)
    converted_points.append([0.0, 0.0])
    visibilities.append(VIS_MISSING)
    return converted_points, visibilities


def convert_side_entry(entry: RawEntry) -> Tuple[List[List[float]], List[int]]:
    if entry.kind == 'manual':
        return convert_manual_side_points(entry.raw_points)
    return convert_auto_side_points(entry.raw_points)


def choose_best(entries: Sequence[RawEntry]) -> RawEntry:
    return max(entries, key=candidate_quality)


def resolve_entries_for_tooth(tooth_id: int,
                              entries: Sequence[RawEntry],
                              summary: dict) -> Dict[str, RawEntry]:
    resolved: Dict[str, RawEntry] = {}
    manual_groups = {'M': [], 'D': []}
    side_groups = {'left': [], 'right': [], 'unknown': []}

    for entry in entries:
        if entry.side_token in {'M', 'D'}:
            manual_groups[entry.side_token].append(entry)
        elif entry.side_token in {'left', 'right'}:
            side_groups[entry.side_token].append(entry)
        else:
            side_groups['unknown'].append(entry)

    for side_name in ('M', 'D'):
        if manual_groups[side_name]:
            if len(manual_groups[side_name]) > 1:
                summary['duplicate_manual_side'] += len(manual_groups[side_name]) - 1
            resolved[side_name] = choose_best(manual_groups[side_name])

    remaining: List[RawEntry] = []
    for bucket in ('left', 'right', 'unknown'):
        remaining.extend(side_groups[bucket])

    if remaining:
        if len(remaining) == 1:
            entry = remaining[0]
            if entry.side_token in {'left', 'right'}:
                mapped_side = image_side_to_md(tooth_id, entry.side_token)
            else:
                if 'M' not in resolved and 'D' in resolved:
                    mapped_side = 'M'
                elif 'D' not in resolved and 'M' in resolved:
                    mapped_side = 'D'
                else:
                    mapped_side = 'M' if mesial_is_image_right(tooth_id) else 'D'
                summary['unknown_side_single_fill'] += 1
            if mapped_side not in resolved:
                resolved[mapped_side] = entry
            else:
                summary['discarded_extra_side'] += 1
        else:
            sorted_entries = sorted(
                remaining,
                key=lambda item: sum(point[0] for point in item.raw_points) /
                max(len(item.raw_points), 1))
            left_entry = sorted_entries[0]
            right_entry = sorted_entries[-1]
            if len(sorted_entries) > 2:
                summary['multi_candidate_tooth'] += len(sorted_entries) - 2

            for mapped_side, entry in (
                    (image_side_to_md(tooth_id, 'left'), left_entry),
                    (image_side_to_md(tooth_id, 'right'), right_entry)):
                if mapped_side not in resolved:
                    resolved[mapped_side] = entry
                else:
                    summary['discarded_extra_side'] += 1

    return resolved


def relative_file_name(dataset_root: Path, image_path: Path) -> str:
    return image_path.relative_to(dataset_root).as_posix()


def build_record(sample_id: int,
                 image_path: Path,
                 traced_image_path: Optional[Path],
                 excel_path: Path,
                 source_canvas_width: Optional[int],
                 source_canvas_height: int,
                 point_offset_x: float,
                 point_offset_y: float,
                 summary: dict) -> ImageRecord:
    with Image.open(image_path) as image:
        width, height = image.size

    workbook = load_workbook(excel_path, data_only=True, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    side_entries: Dict[int, List[RawEntry]] = {
        tooth_id: [] for tooth_id in TARGET_TEETH
    }

    for row_index, row in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True), start=2):
        fallback_tooth_id = None
        if len(row) >= 2:
            fallback_tooth_id = normalize_tooth_id(row[1])
            if fallback_tooth_id is None:
                fallback_tooth_id = normalize_tooth_id(row[0])

        measurement_text = find_measurement_text(row)
        if measurement_text is None:
            continue

        parsed_entries = parse_measurement_entries(
            measurement_text, row_index, fallback_tooth_id, summary)
        if not parsed_entries:
            summary['empty_row_after_parse'] += 1
            continue

        for entry in parsed_entries:
            if entry.tooth_id not in VALID_TOOTH_SET:
                summary['entry_with_invalid_tooth'] += 1
                continue
            if entry.kind == 'unknown':
                summary['unknown_label_format'] += 1
            elif entry.kind == 'null':
                summary['null_label'] += 1
            entry.raw_points = scale_points_to_image(
                entry.raw_points,
                width,
                height,
                source_canvas_width,
                source_canvas_height,
                point_offset_x,
                point_offset_y,
            )
            side_entries[entry.tooth_id].append(entry)

    return ImageRecord(
        sample_id=sample_id,
        image_path=image_path,
        traced_image_path=traced_image_path,
        width=width,
        height=height,
        side_entries=side_entries,
    )


def split_dataset(sample_ids: Sequence[int],
                  train_ratio: float,
                  val_ratio: float,
                  test_ratio: float,
                  seed: int) -> Dict[str, List[int]]:
    total_ratio = train_ratio + val_ratio + test_ratio
    if abs(total_ratio - 1.0) > 1e-6:
        raise ValueError('Split ratios must sum to 1.0.')

    shuffled = list(sample_ids)
    random.Random(seed).shuffle(shuffled)
    num_samples = len(shuffled)
    num_train = int(num_samples * train_ratio)
    num_val = int(num_samples * val_ratio)
    num_test = num_samples - num_train - num_val
    return {
        'train': sorted(shuffled[:num_train]),
        'val': sorted(shuffled[num_train:num_train + num_val]),
        'test': sorted(
            shuffled[num_train + num_val:num_train + num_val + num_test]),
    }
